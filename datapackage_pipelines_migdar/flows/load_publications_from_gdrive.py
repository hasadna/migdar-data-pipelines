from dataflows import Flow, printer, load, dump_to_path, update_resource
from tabulator import Stream
from openpyxl import load_workbook
import tempfile
from datapackage_pipelines_migdar.flows.constants import SEARCH_IMPORT_FIELD_NAMES
from datapackage_pipelines_migdar.flows.common import get_migdar_session


def process_row(row):
    print(row)
    return {k: str(row[k]) if row.get(k) else '' for k in SEARCH_IMPORT_FIELD_NAMES}


def sheet_iterator(first_row, header_row, stream_iter, filename, sheet_name, stream):
    def inner_sheet_iterator():
        if first_row:
            yield process_row(dict(zip(header_row, first_row)))
        for row in stream_iter:
            yield process_row(dict(zip(header_row, row)))

    for i, row in enumerate(inner_sheet_iterator()):
        if not row['migdar_id']:
            if row['title']:
                print('    {}/{}#{}: missing migdar_id'.format(filename, sheet_name, i))
        else:
            yield row
    stream.close()


def load_from_gdrive_files(rows):
    if rows.res.name == 'search_import_index':
        for row_index, row in enumerate(rows, start=1):
            #             if row_index !=5:
            #                 continue
            file_url = f"https://migdar-internal-search.odata.org.il/__data/search_import/{row['name']}"
            print(file_url)
            with tempfile.NamedTemporaryFile('w+b', suffix='.xlsx') as temp_file:
                with get_migdar_session().get(file_url, stream=True) as response:
                    for chunk in response.iter_content():
                        temp_file.write(chunk)
                temp_file.flush()
                wb = load_workbook(temp_file.name)
                for sheet_number, sheet_name in enumerate(wb.sheetnames, start=1):
                    if 'deleted' in sheet_name.strip().lower():
                        continue
                    stream = Stream(temp_file.name, sheet=sheet_name)
                    stream.open()
                    print('#{}.{}/{}: loading sheet'.format(row_index, row['name'], sheet_name))
                    stream_iter = stream.iter()
                    try:
                        first_row = next(stream_iter)
                    except StopIteration:
                        first_row = None
                    if first_row:
                        if 'migdar_id' not in first_row and sheet_number > 1:
                            header_row = first_sheet_header_row
                        else:
                            header_row = first_row
                            first_row = None
                            if sheet_number == 1:
                                first_sheet_header_row = header_row
                        yield from sheet_iterator(first_row, header_row, stream_iter, row['name'], sheet_name, stream)
                    else:
                        for row in stream_iter:
                            pass
    else:
        yield from rows


def flow(*args):
    is_dpp = len(args) > 3
    return Flow(
        load('https://migdar-internal-search.odata.org.il/__data/search_import/index.csv',
             encoding='utf-8', http_session=get_migdar_session()),
        update_resource('index', name='search_import_index', path='search_import_index.csv'),
        load_from_gdrive_files,
        update_resource('search_import_index', name='search_import', path='search_import.csv',
                        schema={'fields': [{'name': n, 'type': 'string'} for n in SEARCH_IMPORT_FIELD_NAMES]},
                        **{'dpp:streaming': True}),
        printer(num_rows=20, tablefmt='plain' if is_dpp else 'html', fields=['migdar_id', 'pubyear', 'title']),
        dump_to_path('data/search_import_from_gdrive')
    )
