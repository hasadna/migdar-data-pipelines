import dataflows as DF
import datetime
import hashlib
import http.client
import json
import multiprocessing
import os
import re
import requests
import socket
import time
import urllib.parse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# The full set of characters RFC 3986 allows in a URL. Anything outside it (a
# space, a Hebrew letter, a quote) terminates the match, which keeps fragments
# like '...#v=onepage' attached instead of clipping the URL in the middle.
URL_RE = re.compile(r'https?://[A-Za-z0-9\-._~:/?#\[\]@!$&\'()*+,;=%]+')
# Punctuation that is legal inside a URL but is almost always sentence-level
# punctuation when it is the last character of a URL found in free text.
TRAILING_PUNCTUATION = '.,;:!?\'")]}>'

# Where the URL was found in the item, as the client sees it on yodaat.org.
FIELD_LABELS = {
    'url': 'קישור למקור',
    'notes': 'אבסטרקט',
    'org_website': 'אתר הארגון',
    'org_facebook': 'פייסבוק',
    'org_email_address': 'דואר אלקטרוני',
    'logo_url': 'לוגו',
    'tagline': 'תיאור קצר',
    'objective': 'מטרות',
    'full_data_source': 'קובץ הנתונים המלא',
    'series.source_url': 'קישור למקור',
    'chart_abstract': 'אבסטרקט',
}

# For every kind of item: which field holds the id/title/authors/abstract the
# client asked for, and which fields may hold links worth checking. Wherever a
# field is translated (foo, foo__en, foo__ar) the untranslated one is the
# Hebrew original, so that is the one we take.
configuration = [
    dict(
        name='publications',
        filename='publications',
        id='migdar_id',
        title='page_title',
        authors='authors',
        abstract='notes',
        link_fields=['url'],
        text_fields=['notes'],
    ),
    dict(
        name='orgs',
        filename='orgs',
        id='entity_id',
        title='org_name',
        authors=None,
        abstract='objective',
        link_fields=['org_website', 'org_facebook', 'org_email_address', 'logo_url'],
        text_fields=['tagline', 'objective'],
    ),
    dict(
        name='datasets',
        filename='out',
        id='doc_id',
        title='chart_title',
        authors='author',
        abstract='chart_abstract',
        link_fields=['full_data_source'],
        json_fields=[('series', 'source_url')],
        text_fields=['chart_abstract'],
    )
]

URL_TEMPLATE='https://api.yodaat.org/data/{name}_in_es/data/{filename}.csv'
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:147.0) Gecko/20100101 Firefox/147.0',
}
TIMEOUT = 15
RETRIES = 3
BACKOFF = 10

STATUS_OK = 'תקין'
STATUS_BROKEN = 'שבור'
STATUS_BLOCKED = 'חסום לבדיקה אוטומטית'
STATUS_TEMPORARY = 'תקלה זמנית'
# Order in which the statuses are presented to the client - the links that are
# certainly dead come first.
STATUS_ORDER = [STATUS_BROKEN, STATUS_TEMPORARY, STATUS_BLOCKED, STATUS_OK]

# A dead link, as opposed to a server that simply refuses to talk to a robot.
BROKEN_CODES = {404, 410}
# Codes that servers hand out to clients they don't like (Cloudflare, gov.il,
# Facebook...). The link itself is usually perfectly fine in a browser.
BLOCKED_CODES = {400, 401, 402, 403, 405, 406, 409, 418, 429, 451, 999}

# Results are cached so that a URL shared by several items is only fetched
# once, and so that a link that failed yesterday can be told apart from one
# that is failing for the first time today.
CACHE_DIR = os.environ.get('BROKEN_LINKS_CACHE_DIR', '.checkpoints/broken-links-cache')
# Shorter than the daily schedule, so every run re-checks every URL exactly
# once no matter how long the run itself takes.
CACHE_TTL = 20 * 3600


def find_urls(value):
    if not isinstance(value, str):
        return []
    urls = (url.rstrip(TRAILING_PUNCTUATION) for url in URL_RE.findall(value))
    return [url for url in urls if len(url) > len('https://')]


def load_json_list(value):
    if isinstance(value, (list, tuple)):
        return value
    try:
        parsed = json.loads(value)
    except (TypeError, ValueError):
        return []
    return parsed if isinstance(parsed, list) else []


def extract_urls(config):
    link_fields = config.get('link_fields') or []
    json_fields = config.get('json_fields') or []
    text_fields = config.get('text_fields') or []

    def func(row):
        found = []
        # Dedicated link fields first, so that a URL appearing both here and in
        # the abstract is reported as the item's 'קישור למקור'.
        for field in link_fields:
            value = row.get(field)
            value = value.strip() if isinstance(value, str) else ''
            # A whole-URL field is taken verbatim - no regexp, nothing to clip.
            urls = [value] if value.startswith('http') else find_urls(value)
            found.extend(dict(url=url, field=field) for url in urls)
        for field, key in json_fields:
            for item in load_json_list(row.get(field)):
                value = item.get(key) if isinstance(item, dict) else None
                value = value.strip() if isinstance(value, str) else ''
                if value.startswith('http'):
                    found.append(dict(url=value, field='%s.%s' % (field, key)))
        for field in text_fields:
            found.extend(dict(url=url, field=field) for url in find_urls(row.get(field)))
        return found

    return func


def unwind():
    def func(rows):
        for row in rows:
            seen = set()
            for entry in row['urls']:
                url = entry['url']
                if url in seen:
                    continue
                seen.add(url)
                # A fresh dict per URL: the row is handed to another process by
                # DF.parallelize, which may pickle it after we've moved on.
                out = dict(row)
                out['url'] = url
                out['url_source'] = FIELD_LABELS.get(entry['field'], entry['field'])
                yield out
    return func


def cache_filename(url):
    digest = hashlib.sha1(url.encode('utf-8')).hexdigest()
    return os.path.join(CACHE_DIR, digest + '.json')


def cache_get(url):
    try:
        with open(cache_filename(url), encoding='utf-8') as cached:
            entry = json.load(cached)
    except (IOError, OSError, ValueError):
        return None
    if not isinstance(entry, dict) or entry.get('url') != url:
        return None
    return entry


def cache_set(url, entry):
    filename = cache_filename(url)
    # Written from 16 processes at once - rename is atomic, a partial write
    # never becomes visible under the real name.
    temp = '%s.%s.tmp' % (filename, os.getpid())
    try:
        os.makedirs(CACHE_DIR, exist_ok=True)
        with open(temp, 'w', encoding='utf-8') as out:
            json.dump(entry, out, ensure_ascii=False)
        os.replace(temp, filename)
    except (IOError, OSError):
        pass


def probe(url, method):
    try:
        resp = requests.request(method, url, allow_redirects=True,
                                headers=HEADERS, timeout=TIMEOUT, stream=True)
        try:
            # Not every server sends a reason phrase, and 'שגיאה: 404' with
            # nothing after the colon reads like a bug to whoever gets the file.
            reason = resp.reason or http.client.responses.get(resp.status_code, '')
            return resp.status_code, reason
        finally:
            resp.close()
    except Exception as e:
        return None, str(e.__class__.__name__)


def host_resolves(url):
    host = urllib.parse.urlsplit(url).hostname
    if not host:
        return False
    try:
        socket.getaddrinfo(host, None)
    except socket.gaierror:
        return False
    except Exception:
        # Anything else (a timeout in the resolver, say) is not evidence that
        # the domain is gone.
        return True
    return True


def classify(code, reason):
    if code is None:
        return STATUS_TEMPORARY, reason
    detail = '%s: %s' % (code, reason)
    if code < 300:
        return STATUS_OK, None
    if code in BROKEN_CODES:
        return STATUS_BROKEN, detail
    if code in BLOCKED_CODES:
        return STATUS_BLOCKED, detail
    if code >= 500:
        return STATUS_TEMPORARY, detail
    return STATUS_BROKEN, detail


def check_url(url):
    code, reason = probe(url, 'HEAD')
    backoff = BACKOFF
    for _ in range(RETRIES - 1):
        if code != 429:
            break
        time.sleep(backoff)
        backoff *= 2
        code, reason = probe(url, 'HEAD')
    status, error = classify(code, reason)
    # Plenty of servers reject HEAD outright, or answer it with a 403 they
    # would never give a browser. A GET is the tiebreaker; we never read the
    # body, so it costs about the same.
    if status == STATUS_BLOCKED or (code is not None and code >= 500):
        time.sleep(1)
        status, error = classify(*probe(url, 'GET'))
    if status == STATUS_TEMPORARY and code is None and not host_resolves(url):
        # The domain itself no longer resolves - that is a dead link, not a
        # server having a bad morning.
        return STATUS_BROKEN, 'DNS: no such host'
    return status, error


def check_broken():
    def func(row):
        url = row['url']
        now = time.time()
        entry = cache_get(url)
        if entry is None or now - entry.get('checked_at', 0) > CACHE_TTL:
            print('%s:CHECK:%s' % (datetime.datetime.now().isoformat(), url))
            status, error = check_url(url)
            streak = 0 if entry is None else entry.get('fail_streak', 0)
            entry = dict(
                url=url,
                status=status,
                error=error,
                checked_at=now,
                fail_streak=0 if status == STATUS_OK else streak + 1,
            )
            cache_set(url, entry)
            if status != STATUS_OK:
                print('%s:%s:%s: %s' % (datetime.datetime.now().isoformat(),
                                        status, url, error))
            time.sleep(1)
        row['status'] = entry['status']
        row['error'] = entry['error']
        row['fail_streak'] = entry['fail_streak']
        row['checked_at'] = datetime.datetime.fromtimestamp(entry['checked_at'])\
                                             .replace(microsecond=0)
    return func


def get_field(field_name):
    def func(r):
        if field_name is None:
            return ''
        if field_name not in r:
            print('ERRRR, missing field %s in %r' % (field_name, r))
            return ''
        value = r[field_name]
        return '' if value is None else str(value)
    return func


def broken_links_flow(limit_rows=None):
    # A limited run gets its own checkpoint, so that a quick sample never ends
    # up being served from - or serving - the cache of a full run.
    checkpoint_name = 'broken_links' if limit_rows is None else 'broken_links_%d' % limit_rows
    DF.Flow(
        *[
            DF.Flow(
                DF.load(URL_TEMPLATE.format(**c), name=c['name'], limit_rows=limit_rows),
                DF.add_field('__name', 'string', c['name'], resources=c['name']),
                DF.add_field('__title', 'string', get_field(c['title']), resources=c['name']),
            )
            for c in configuration
        ],
        DF.checkpoint(checkpoint_name),
    ).process()
    return DF.Flow(
        DF.checkpoint(checkpoint_name),
        *[
            DF.Flow(
                DF.add_field('__id', 'string', get_field(c['id']), resources=c['name']),
                DF.add_field('__authors', 'string', get_field(c.get('authors')), resources=c['name']),
                DF.add_field('__abstract', 'string', get_field(c.get('abstract')), resources=c['name']),
                DF.add_field('__urls', 'array', extract_urls(c), resources=c['name']),
            )
            for c in configuration
        ],
        DF.add_field('link', 'string', lambda r: 'https://yodaat.org/item/{doc_id}'.format(**r)),
        DF.concatenate(dict(
            item_id=['__id'],
            title=['__title'],
            authors=['__authors'],
            abstract=['__abstract'],
            link=[],
            urls=['__urls'],
            name=['__name'],
        )),
        DF.add_field('url', 'string'),
        DF.add_field('url_source', 'string'),
        DF.add_field('status', 'string'),
        DF.add_field('error', 'string'),
        DF.add_field('fail_streak', 'integer'),
        DF.add_field('checked_at', 'datetime'),
        unwind(),
        DF.delete_fields(['urls']),
        DF.parallelize(check_broken(), 16),
        DF.filter_rows(lambda r: r['status'] != STATUS_OK),
    )


# The six columns the client asked for, in the order they asked for them,
# followed by the context needed to act on each row.
COLUMNS = [
    ('item_id', 'מס\' זיהוי', 12),
    ('title', 'כותרת הפריט', 55),
    ('authors', 'מחברים', 25),
    ('abstract', 'אבסטרקט', 70),
    ('link', 'קישור לפריט ביודעת', 38),
    ('url', 'הקישור השבור', 55),
    ('url_source', 'מיקום הקישור בפריט', 18),
    ('name', 'סוג הפריט', 12),
    ('status', 'סטטוס', 20),
    ('error', 'פירוט השגיאה', 22),
    ('fail_streak', 'בדיקות כושלות ברצף', 10),
    ('checked_at', 'תאריך הבדיקה', 18),
]
LINK_COLUMNS = {'link', 'url'}
# Control characters openpyxl refuses to write into a cell.
ILLEGAL_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
MAX_CELL_LENGTH = 32000


def clean_cell(value):
    if value is None:
        return ''
    if isinstance(value, datetime.datetime):
        return value.replace(tzinfo=None, microsecond=0)
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARS.sub(' ', value)[:MAX_CELL_LENGTH]


def convert_to_xlsx(out_path, source='data/broken_links/datapackage.json'):
    def func():
        rows = DF.Flow(DF.load(source)).results()[0][0]
        rows.sort(key=lambda r: (
            STATUS_ORDER.index(r['status']) if r['status'] in STATUS_ORDER else len(STATUS_ORDER),
            r.get('name') or '',
            r.get('item_id') or '',
        ))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'קישורים שבורים'
        sheet.sheet_view.rightToLeft = True

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4F6228')
        link_font = Font(color='0563C1', underline='single')
        sheet.append([title for _, title, _ in COLUMNS])
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in rows:
            sheet.append([clean_cell(row.get(field)) for field, _, _ in COLUMNS])
            written = sheet[sheet.max_row]
            for cell, (field, _, _) in zip(written, COLUMNS):
                cell.alignment = Alignment(vertical='top', wrap_text=field in ('title', 'abstract'))
                if field in LINK_COLUMNS and isinstance(cell.value, str) and cell.value.startswith('http'):
                    try:
                        cell.hyperlink = cell.value
                        cell.font = link_font
                    except ValueError:
                        pass
                elif field == 'checked_at' and cell.value:
                    cell.number_format = 'YYYY-MM-DD HH:MM'

        for index, (_, _, width) in enumerate(COLUMNS, start=1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions

        os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
        workbook.save(out_path)
        print('WROTE %d ROWS TO %s' % (len(rows), out_path))

    return func


def flow(*_):
    return DF.Flow(
        broken_links_flow(),
        DF.update_resource(-1, **{'dpp:streaming': True}),
        DF.printer(),
        DF.dump_to_path('data/broken_links'),
        DF.finalizer(convert_to_xlsx('data/broken_links.xlsx')),
    )

if __name__ == '__main__':
    # DF.parallelize hands a closure to each worker process, which survives
    # only under the 'fork' start method. That is the default wherever the
    # pipeline actually runs; on macOS it is 'spawn', and the workers never
    # come up, so a direct run hangs waiting for results that never arrive.
    multiprocessing.set_start_method('fork')
    # Run directly, this is a smoke test - the pipeline imports flow() instead.
    DF.Flow(
        broken_links_flow(limit_rows=10),
        DF.printer(),
        DF.dump_to_path('data/broken_links'),
        DF.finalizer(convert_to_xlsx('data/broken_links.xlsx')),
    ).process()
