import os
import subprocess

from openpyxl import load_workbook

import dataflows as DF

SCREENSHOT = os.path.join(os.path.dirname(__file__), 'node', 'screenshot.js')
XLSX_TEMPLATE = os.path.join(os.path.dirname(__file__), 'template.xlsx')
EXCEL_METADATA = [
    ('B1', 'chart_title'),
    ('B2', 'chart_abstract'),
    ('B3', 'units'),
    ('B4', 'source_description'),
    ('B5', 'source_detail_description'),
    ('B6', 'extrapulation_years'),
]
EXCEL_SERIES_ROW = 9


def write_excel():

    def func(row):
        wb = load_workbook(XLSX_TEMPLATE)
        ws = wb.active
        for cell, field in EXCEL_METADATA:
            ws[cell] = row.get(field, row['series'][0].get(field, ''))
        source_range = set()
        for x, series in enumerate(row['series']):
            ws.cell(EXCEL_SERIES_ROW, x+2, series['gender'])
            source_range.update(d['x'] for d in series['dataset'])
        source_range = dict((v, i + 1) for i, v in enumerate(sorted(source_range)))
        for value, y in source_range.items():
            try:
                ws.cell(EXCEL_SERIES_ROW + y, 1, int(value))
            except:
                ws.cell(EXCEL_SERIES_ROW + y, 1, value)
        for x, series in enumerate(row['series']):
            for value in series['dataset']:
                y = source_range[value['x']]
                ws.cell(EXCEL_SERIES_ROW + y, x+2, value['y'])
        outpath = os.path.join('data', row['doc_id'] + '.xlsx')
        wb.save(filename=outpath)

    return func


def do_screenshot():
    def func(rows):
        for row in rows:
            for lang in ['', 'ar/', 'en/']:
                doc_id = row['doc_id']
                for suffix in ['', '-share']:
                    url = f'https://yodaat.org/{lang}card{suffix}/{doc_id}'
                    outpath = os.path.join('data',
                                           lang + os.path.dirname(doc_id))
                    os.makedirs(outpath, exist_ok=True)
                    outpath = os.path.join('data',
                                           lang + doc_id + suffix + '.png')
                    subprocess.call(['node', SCREENSHOT, url,
                                     outpath, '.card'])
            yield row
    return func


def flow(*_, path='data/datasets_in_es'):
    return DF.Flow(
        DF.load('{}/datapackage.json'.format(path)),
        do_screenshot(),
        write_excel(),
        DF.update_resource(-1, **{'dpp:streaming': True})
    )


if __name__ == '__main__':
    flow(path='https://api.yodaat.org/data/datasets_in_es')\
        .process()
